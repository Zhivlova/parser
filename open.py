import openpyxl
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer

wb = load_workbook('excel.xlsm')
sheet = wb.active
sheet = wb.worksheets[1]

for item in range(1, 31):
    # print(sheet['A' + str(item)].value, sheet['B' + str(item)].value, sheet['C' + str(item)].value, sheet['D' + str(item)].value)
    print(sheet['M' + str(item)].value)
# for row in sheet.iter_rows(min_row=1, max_row=12, min_col=0, max_col=4):
#     for cell in row:
#         print(cell.value, end='')
#     print()


# control_actions_of_the_model = {}
#
cells = sheet['A3': 'D12']
for name, designation, basic_equilibrium, new_equilibrium in cells:
#     name = name.value
#     print(name)
    print(name.value, designation.value, basic_equilibrium.value, new_equilibrium.value)

# for row in range(3, 13):
#     name = sheet[row][0].value
#     designation = sheet[row][1].value
#     basic_equilibrium = sheet[row][2].value
#     new_equilibrium = sheet[row][3].value
#     print(row, name, designation, basic_equilibrium, new_equilibrium)



