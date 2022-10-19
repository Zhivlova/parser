import os
from openpyxl import Workbook, load_workbook
import csv

import pandas as pd

mydir = '/Users/natalazivlova/Desktop/parser/28.09/'
myfile = 'рейтинг 2021 корр.xlsx'
file = os.path.join(mydir, myfile)

wb = load_workbook(file, data_only=True)

print(f'Структура эксель файла {wb.sheetnames}')
print('------------------------------------------------------------------------------')

sheet = wb.worksheets[0]

sonko = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    sonko['subject'].append(subject)
    scores = sheet[row][2].value
    sonko['scores'].append(scores)
print(sonko)

lines1 = pd.DataFrame(sonko)

social_entrepreneurship = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    social_entrepreneurship['subject'].append(subject)
    scores = sheet[row][3].value
    social_entrepreneurship['scores'].append(scores)
print(social_entrepreneurship)

sonco_taxes = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    sonco_taxes['subject'].append(subject)
    scores = sheet[row][4].value
    sonco_taxes['scores'].append(scores)
print(sonco_taxes)

taxes_donations = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    taxes_donations['subject'].append(subject)
    scores = sheet[row][5].value
    taxes_donations['scores'].append(scores)
print(taxes_donations)

mun_prog_sonko = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    mun_prog_sonko['subject'].append(subject)
    scores = sheet[row][6].value
    mun_prog_sonko['scores'].append(scores)
print(mun_prog_sonko)

mun_prog_sopred = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    mun_prog_sopred['subject'].append(subject)
    scores = sheet[row][7].value
    mun_prog_sopred['scores'].append(scores)
print(mun_prog_sopred)

infrastructure = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    infrastructure['subject'].append(subject)
    scores = sheet[row][8].value
    infrastructure['scores'].append(scores)
print(infrastructure)

workers = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    workers['subject'].append(subject)
    scores = sheet[row][9].value
    workers['scores'].append(scores)
print(workers)

social_entrepreneurship_support = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    social_entrepreneurship_support['subject'].append(subject)
    scores = sheet[row][10].value
    social_entrepreneurship_support['scores'].append(scores)
print(social_entrepreneurship_support)

soc_order = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    soc_order['subject'].append(subject)
    scores = sheet[row][11].value
    soc_order['scores'].append(scores)
print(soc_order)

personal_financing = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    personal_financing['subject'].append(subject)
    scores = sheet[row][12].value
    personal_financing['scores'].append(scores)
print(personal_financing)

gchp_not_public = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    gchp_not_public['subject'].append(subject)
    scores = sheet[row][13].value
    gchp_not_public['scores'].append(scores)
print(gchp_not_public)

website = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    website['subject'].append(subject)
    scores = sheet[row][14].value
    website['scores'].append(scores)
print(website)

clicks = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    clicks['subject'].append(subject)
    scores = sheet[row][15].value
    clicks['scores'].append(scores)
print(clicks)

public_licenses = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    public_licenses['subject'].append(subject)
    scores = sheet[row][16].value
    public_licenses['scores'].append(scores)
print(public_licenses)

open_data = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    open_data['subject'].append(subject)
    scores = sheet[row][17].value
    open_data['scores'].append(scores)
print(open_data)

children = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    children['subject'].append(subject)
    scores = sheet[row][18].value
    children['scores'].append(scores)
print(children)

medical_organizations = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    medical_organizations['subject'].append(subject)
    scores = sheet[row][19].value
    medical_organizations['scores'].append(scores)
print(medical_organizations)

suppliers = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    suppliers['subject'].append(subject)
    scores = sheet[row][20].value
    suppliers['scores'].append(scores)
print(suppliers)

culture = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    culture['subject'].append(subject)
    scores = sheet[row][21].value
    culture['scores'].append(scores)
print(culture)

final_results = {"subject": [], "scores": []}
for row in range(2, 87):
    subject = sheet[row][1].value
    final_results['subject'].append(subject)
    scores = sheet[row][22].value
    final_results['scores'].append(scores)
print(final_results)



with open('scores.csv', 'w') as file:
    writer = csv.writer(file)
    writer.writerow(["Субъект", "Баллы"])
# for i in final_results:
#     subject = i[0]
#     scores = i[1]
#
#     with open('scores.csv', 'w') as file:
#         writer = csv.writer(file)
#         writer.writerow(
#             i
#         )