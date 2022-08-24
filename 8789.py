from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
import openpyxl
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer

wb = load_workbook('excel.xlsm')
# new_range = DefinedName('newrange', attr_text='Sheet!$C$1:$A$5')
# wb.defined_names.append(new_range)
# print(wb.defined_names)
print(wb.defined_names['name'])
#
# cells = []
# for title, coord in wb.defined_names:
#     ws = wb[title]
#     cells.append(ws[coord])
# print(cells)

# my_range = wb.defined_names['P_SDP_0']
# dests = my_range.destinations
# print(dests)
# cells = []
# for title, coord in dests:
#     ws = wb[title]
#     cells.append(ws[coord])
# print(cells)

# my_range = wb.defined_names['P_SDP_0']
# dests = my_range.destinations
#
# cells = []
# for title, coord in dests:
#     ws = wb[title]
#     cells.append(ws[coord])
# print([cell.value for cell in cells])